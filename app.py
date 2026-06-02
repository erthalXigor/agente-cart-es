import streamlit as st
import anthropic
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
import os

st.set_page_config(
    page_title="Agente de Cartões Brasil 2026",
    page_icon="💳",
    layout="centered",
)

# ── CSS para botões de seleção ───────────────────────────────────────────────
st.markdown("""
<style>
div[data-testid="stHorizontalBlock"] button {
    border-radius: 20px !important;
}
</style>
""", unsafe_allow_html=True)

@st.cache_data(show_spinner="Carregando base de conhecimento...")
def carregar_base():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    docx_path = os.path.join(base_dir, "referencia_cartoes_agente_2026.docx")
    paragrafos = []
    with zipfile.ZipFile(docx_path, "r") as z:
        with z.open("word/document.xml") as f:
            tree = ET.parse(f)
            ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
            for para in tree.findall(".//w:p", ns):
                texto = "".join(r.text or "" for r in para.findall(".//w:t", ns))
                if texto.strip():
                    paragrafos.append(texto.strip())
    doc_texto = "\n".join(paragrafos)

    xlsx_path = os.path.join(base_dir, "cartoes_credito_brasil_2026_v2.xlsx")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    abas = []
    for nome in wb.sheetnames:
        ws = wb[nome]
        linhas = []
        for row in ws.iter_rows(values_only=True):
            linha = "\t".join(str(c) if c is not None else "" for c in row)
            if linha.strip():
                linhas.append(linha)
        abas.append(f"=== ABA: {nome} ===\n" + "\n".join(linhas))
    xlsx_texto = "\n\n".join(abas)

    return f"""Você é o Agente de Cartões de Crédito Brasil 2026.
Sua base de conhecimento contém dados detalhados sobre cartões de crédito brasileiros, programas de pontos, perfis de usuário e estratégias de recomendação.

=== DOCUMENTO DE REFERÊNCIA ===
{doc_texto}

=== PLANILHA DE CARTÕES ===
{xlsx_texto}

Regras:
- Seja direto, claro e objetivo.
- Use formatação markdown nas respostas.
- Baseie todas as respostas EXCLUSIVAMENTE na base de conhecimento acima.
- Se uma informação não estiver na base, diga que não tem essa informação.
"""


def chamar_claude(sistema: str, mensagem: str) -> str:
    api_key = os.environ.get("ANTHROPIC_API_KEY", "") or st.session_state.get("api_key", "")
    client = anthropic.Anthropic(api_key=api_key)
    resposta = ""
    with client.messages.stream(
        model="claude-sonnet-4-6",
        max_tokens=2048,
        system=[{"type": "text", "text": sistema, "cache_control": {"type": "ephemeral"}}],
        messages=[{"role": "user", "content": mensagem}],
    ) as stream:
        for texto in stream.text_stream:
            resposta += texto
    return resposta


def botoes_selecao_unica(label, opcoes, key):
    """Seção colapsável com botões de seleção única."""
    with st.expander(label, expanded=False):
        selecionado = st.session_state.get(key, opcoes[0])
        cols = st.columns(2)
        for i, op in enumerate(opcoes):
            tipo = "primary" if selecionado == op else "secondary"
            if cols[i % 2].button(op, key=f"{key}_{i}", type=tipo, use_container_width=True):
                st.session_state[key] = op
                st.rerun()
    return st.session_state.get(key, opcoes[0])


def botoes_selecao_multipla(label, opcoes, key):
    """Seção colapsável com botões de seleção múltipla."""
    with st.expander(label, expanded=False):
        selecionados = st.session_state.get(key, [])
        cols = st.columns(2)
        for i, op in enumerate(opcoes):
            marcado = op in selecionados
            tipo = "primary" if marcado else "secondary"
            if cols[i % 2].button(op, key=f"{key}_{i}", type=tipo, use_container_width=True):
                if marcado:
                    selecionados = [x for x in selecionados if x != op]
                else:
                    selecionados = selecionados + [op]
                st.session_state[key] = selecionados
                st.rerun()
    return st.session_state.get(key, [])


# ── Estado da sessão ─────────────────────────────────────────────────────────
if "api_key" not in st.session_state:
    st.session_state.api_key = os.environ.get("ANTHROPIC_API_KEY", "")
if "modo" not in st.session_state:
    st.session_state.modo = None
if "resultado" not in st.session_state:
    st.session_state.resultado = ""

# ── Header ───────────────────────────────────────────────────────────────────
st.title("💳 Agente de Cartões de Crédito")
st.caption("Brasil 2026 · Powered by Claude")

# ── API Key ───────────────────────────────────────────────────────────────────
# Tenta pegar do ambiente novamente (Railway injeta após inicialização)
if not st.session_state.api_key:
    st.session_state.api_key = os.environ.get("ANTHROPIC_API_KEY", "")

if not st.session_state.api_key:
    with st.expander("🔑 Configurar chave de API", expanded=True):
        chave = st.text_input("ANTHROPIC_API_KEY", type="password", placeholder="sk-ant-...")
        if st.button("Salvar"):
            st.session_state.api_key = chave
            st.rerun()
    st.stop()

try:
    sistema = carregar_base()
except FileNotFoundError as e:
    st.error(f"Arquivo não encontrado: {e}")
    st.stop()

# ── Menu principal ────────────────────────────────────────────────────────────
st.divider()
col1, col2, col3, col4 = st.columns(4)
with col1:
    if st.button("🎯 Recomendar", use_container_width=True):
        st.session_state.modo = "recomendar"
        st.session_state.resultado = ""
with col2:
    if st.button("⚖️ Comparar", use_container_width=True):
        st.session_state.modo = "comparar"
        st.session_state.resultado = ""
with col3:
    if st.button("💰 Vale a anuidade?", use_container_width=True):
        st.session_state.modo = "anuidade"
        st.session_state.resultado = ""
with col4:
    if st.button("📋 Explicar cartão", use_container_width=True):
        st.session_state.modo = "explicar"
        st.session_state.resultado = ""

st.divider()

# ── Modo: Recomendar ──────────────────────────────────────────────────────────
if st.session_state.modo == "recomendar":
    st.subheader("🎯 Recomendar cartão para o meu perfil")

    RENDAS = [
        "Até R$ 1.000", "R$ 1.000 – R$ 3.000", "R$ 3.000 – R$ 5.000",
        "R$ 5.000 – R$ 10.000", "R$ 10.000 – R$ 20.000", "Acima de R$ 20.000"
    ]
    OBJETIVOS = [
        "Milhas LATAM", "Milhas Azul", "Milhas Smiles (GOL)",
        "Cashback", "Lounge / Sala VIP", "Sem anuidade / custo zero"
    ]
    BANCOS = [
        "XP Investimentos", "BTG Pactual", "Banco Inter", "Banco do Brasil",
        "Bradesco", "Santander", "Itaú", "Nubank", "Caixa Econômica Federal",
        "C6 Bank", "Nenhum"
    ]

    renda = botoes_selecao_unica("💰 Renda mensal aproximada", RENDAS, "rec_renda")
    st.caption(f"Selecionado: **{renda}**")

    objetivos = botoes_selecao_multipla("🎯 Objetivos (pode escolher mais de um)", OBJETIVOS, "rec_objetivos")
    if objetivos:
        st.caption(f"Selecionados: **{', '.join(objetivos)}**")

    bancos = botoes_selecao_multipla("🏦 Investimentos em bancos (pode escolher mais de um)", BANCOS, "rec_bancos")
    if bancos:
        st.caption(f"Selecionados: **{', '.join(bancos)}**")

    # Valor investido — só para bancos selecionados (exceto "Nenhum")
    bancos_com_valor = [b for b in bancos if b != "Nenhum"]
    valores_investidos = {}
    if bancos_com_valor:
        st.markdown("**Valor investido por banco:**")
        for banco in bancos_com_valor:
            valores_investidos[banco] = st.number_input(
                f"{banco} (R$)", min_value=0, step=10000, value=0, key=f"val_{banco}"
            )

    viaja = st.radio("✈️ Viaja com frequência?", ["Sim", "Não"], horizontal=True)
    gasto = st.text_input("💳 Gasto médio mensal no cartão (R$)", placeholder="ex: 5000")

    if st.button("🎯 Recomendar", type="primary", use_container_width=True):
        inv_texto = "Nenhum"
        if bancos_com_valor:
            partes = []
            for banco in bancos_com_valor:
                val = valores_investidos.get(banco, 0)
                partes.append(f"{banco}: R$ {val:,.0f}" if val > 0 else banco)
            inv_texto = ", ".join(partes)

        prompt = (
            f"Perfil do usuário:\n"
            f"- Renda mensal: {renda}\n"
            f"- Objetivos: {', '.join(objetivos) if objetivos else 'Não especificado'}\n"
            f"- Investimentos: {inv_texto}\n"
            f"- Viaja frequentemente: {viaja}\n"
            f"- Gasto médio mensal no cartão: R$ {gasto or 'Não informado'}\n\n"
            f"Com base nesse perfil, recomende até 3 cartões de crédito. "
            f"Para cada um, explique por que faz sentido, qual a anuidade e se há algum alerta para 2026. "
            f"Leve em conta os investimentos bancários para identificar cartões exclusivos por relacionamento."
        )
        with st.spinner("Analisando seu perfil..."):
            st.session_state.resultado = chamar_claude(sistema, prompt)

# ── Modo: Comparar ────────────────────────────────────────────────────────────
elif st.session_state.modo == "comparar":
    st.subheader("⚖️ Comparar dois cartões")
    with st.form("form_comparar"):
        cartao1 = st.text_input("Primeiro cartão", placeholder="ex: Nubank Ultravioleta")
        cartao2 = st.text_input("Segundo cartão", placeholder="ex: C6 Carbon")
        submitted = st.form_submit_button("Comparar", type="primary", use_container_width=True)
    if submitted and cartao1 and cartao2:
        prompt = (
            f"Compare os cartões '{cartao1}' e '{cartao2}' lado a lado. "
            f"Use uma tabela markdown com os seguintes critérios: anuidade, programa de pontos, "
            f"taxa de conversão, benefícios principais, lounge, seguros, e para quem é mais indicado."
        )
        with st.spinner("Comparando cartões..."):
            st.session_state.resultado = chamar_claude(sistema, prompt)

# ── Modo: Vale a anuidade ─────────────────────────────────────────────────────
elif st.session_state.modo == "anuidade":
    st.subheader("💰 Vale a pena pagar a anuidade?")
    with st.form("form_anuidade"):
        cartao = st.text_input("Qual cartão?", placeholder="ex: Itaú Personnalité Visa Infinite")
        gasto_mensal = st.number_input("Gasto médio mensal no cartão (R$)", min_value=0, value=3000, step=500)
        uso_lounge = st.radio("Usa salas VIP em aeroportos?", ["Sim", "Não", "Às vezes"], horizontal=True)
        submitted = st.form_submit_button("Calcular", type="primary", use_container_width=True)
    if submitted and cartao:
        prompt = (
            f"O usuário quer saber se vale a pena pagar a anuidade do '{cartao}'.\n"
            f"- Gasto médio mensal: R$ {gasto_mensal:,.0f}\n"
            f"- Uso de sala VIP: {uso_lounge}\n\n"
            f"Faça o cálculo com números reais: quantos pontos/milhas ele acumula por ano, "
            f"quanto isso vale em reais, some o valor dos benefícios que ele usa, "
            f"e compare com o custo da anuidade. Conclua se vale ou não."
        )
        with st.spinner("Calculando..."):
            st.session_state.resultado = chamar_claude(sistema, prompt)

# ── Modo: Explicar ────────────────────────────────────────────────────────────
elif st.session_state.modo == "explicar":
    st.subheader("📋 Explicar um cartão específico")
    with st.form("form_explicar"):
        cartao = st.text_input("Nome do cartão", placeholder="ex: XP Visa Infinite")
        submitted = st.form_submit_button("Explicar", type="primary", use_container_width=True)
    if submitted and cartao:
        prompt = (
            f"Explique tudo sobre o cartão '{cartao}': anuidade, como isentar, programa de pontos, "
            f"taxa de conversão por categoria, benefícios (lounge, seguros, concierge), "
            f"renda mínima exigida, para qual perfil é indicado e alertas importantes para 2026."
        )
        with st.spinner("Buscando informações..."):
            st.session_state.resultado = chamar_claude(sistema, prompt)

# ── Resultado ─────────────────────────────────────────────────────────────────
if st.session_state.resultado:
    st.divider()
    st.markdown(st.session_state.resultado)
    if st.button("🔄 Nova consulta"):
        st.session_state.resultado = ""
        st.session_state.modo = None
        st.rerun()
