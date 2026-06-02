#!/usr/bin/env python3
"""
Agente de Cartões de Crédito Brasil 2026
"""

import os
import sys
import zipfile
import xml.etree.ElementTree as ET

try:
    import anthropic
except ImportError:
    print("Erro: execute  pip3 install anthropic python-docx openpyxl")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Erro: execute  pip3 install openpyxl")
    sys.exit(1)


# ── Leitura dos arquivos ────────────────────────────────────────────────────


def ler_docx(caminho: str) -> str:
    with zipfile.ZipFile(caminho) as z:
        with z.open("word/document.xml") as f:
            root = ET.parse(f).getroot()
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    linhas = []
    for p in root.findall(".//w:p", ns):
        texto = "".join(t.text for t in p.findall(".//w:t", ns) if t.text)
        if texto.strip():
            linhas.append(texto)
    return "\n".join(linhas)


def ler_xlsx(caminho: str) -> str:
    wb = openpyxl.load_workbook(caminho)
    blocos = []
    for nome in wb.sheetnames:
        ws = wb[nome]
        linhas = [f"\n=== {nome} ==="]
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                linhas.append("\t".join("" if c is None else str(c) for c in row))
        blocos.append("\n".join(linhas))
    return "\n".join(blocos)


def montar_sistema(doc: str, xlsx: str) -> str:
    return f"""Você é um especialista em cartões de crédito brasileiros em 2026.
Sua base de conhecimento está abaixo — use-a como fonte autoritativa.

=== DOCUMENTO DE REFERÊNCIA ===
{doc}

=== PLANILHA DE CARTÕES (33 cartões, 8 abas) ===
{xlsx}

REGRAS OBRIGATÓRIAS:
- Nunca recomende mais de 3 cartões por vez
- Sempre explique como isentar a anuidade de cada cartão recomendado
- Alerte sobre mudanças de 2026 quando relevante
- Responda em português brasileiro
- Use formatação clara com tópicos e separadores
- Seja direto e objetivo — sem enrolação"""


# ── Chamada à API com streaming e prompt caching ───────────────────────────


def claude(client: anthropic.Anthropic, sistema: str, prompt: str) -> None:
    print()
    with client.messages.stream(
        model="claude-sonnet-4-6",
        max_tokens=2048,
        system=[
            {
                "type": "text",
                "text": sistema,
                "cache_control": {"type": "ephemeral"},  # cache do sistema entre chamadas
            }
        ],
        messages=[{"role": "user", "content": prompt}],
    ) as stream:
        for chunk in stream.text_stream:
            print(chunk, end="", flush=True)
    print("\n")


# ── Utilitários de UI ───────────────────────────────────────────────────────


def hr(char: str = "─", width: int = 54) -> None:
    print(char * width)


def perguntar(msg: str) -> str:
    return input(msg).strip()


def menu_objetivo() -> str:
    print("\n  Objetivo principal:")
    opcoes = {
        "1": "milhas LATAM",
        "2": "milhas Azul",
        "3": "milhas Smiles (GOL)",
        "4": "cashback",
        "5": "lounge / sala VIP em aeroporto",
        "6": "sem anuidade / custo zero",
    }
    for k, v in opcoes.items():
        print(f"  [{k}] {v.capitalize()}")
    return opcoes.get(perguntar("  Escolha: "), "milhas em geral")


def menu_banco() -> str:
    print("\n  Investimentos em algum banco específico?")
    opcoes = {
        "1": "XP Investimentos",
        "2": "BTG Pactual",
        "3": "Banco Inter",
        "4": "Banco do Brasil",
        "5": "Bradesco",
        "6": "Santander",
        "7": "Nenhum banco específico",
    }
    for k, v in opcoes.items():
        print(f"  [{k}] {v}")
    return opcoes.get(perguntar("  Escolha: "), "nenhum banco")


# ── Opções do menu ─────────────────────────────────────────────────────────


def op_recomendar(client: anthropic.Anthropic, sistema: str) -> None:
    hr()
    print("  RECOMENDAÇÃO POR PERFIL")
    hr()

    renda = perguntar("\n  Renda mensal aproximada: R$ ")
    objetivo = menu_objetivo()
    banco = menu_banco()
    viaja_raw = perguntar("\n  Viaja com frequência? (s/n): ").lower()
    viaja = "sim, viaja frequentemente" if viaja_raw in ("s", "sim") else "não viaja muito"

    prompt = f"""Perfil do usuário:
- Renda mensal: R$ {renda}
- Objetivo principal: {objetivo}
- Investimentos em: {banco}
- Frequência de viagens: {viaja}

Com base nesse perfil, recomende até 3 cartões. Para cada um:
1. Nome e motivo objetivo da indicação para esse perfil
2. Como isentar a anuidade (valor e condição exata)
3. Alerta de 2026 relevante, se houver

Termine com uma recomendação final destacada ("Minha escolha principal para esse perfil: …")."""

    claude(client, sistema, prompt)


def op_comparar(client: anthropic.Anthropic, sistema: str) -> None:
    hr()
    print("  COMPARAÇÃO LADO A LADO")
    hr()

    cartao1 = perguntar("\n  Nome do 1º cartão: ")
    cartao2 = perguntar("  Nome do 2º cartão: ")

    prompt = f"""Compare "{cartao1}" vs "{cartao2}" nas seguintes dimensões:

| Dimensão              | {cartao1} | {cartao2} |
|-----------------------|-----------|-----------|
| Anuidade anual        |           |           |
| Como isentar          |           |           |
| Pontos/cashback       |           |           |
| Lounge / Sala VIP     |           |           |
| Programa de milhas    |           |           |
| Renda / requisito     |           |           |
| Perfil ideal          |           |           |
| Alertas 2026          |           |           |

Após a tabela, indique em um parágrafo: para qual perfil cada cartão faz mais sentido.
Se algum cartão não estiver na base de conhecimento, informe e sugira o mais parecido."""

    claude(client, sistema, prompt)


def op_vale_anuidade(client: anthropic.Anthropic, sistema: str) -> None:
    hr()
    print("  VALE PAGAR A ANUIDADE?")
    hr()

    cartao = perguntar("\n  Nome do cartão: ")
    gasto = perguntar("  Gasto mensal estimado nesse cartão: R$ ")

    prompt = f"""O usuário quer saber se vale pagar a anuidade do cartão "{cartao}" gastando R$ {gasto}/mês.

Analise com números reais:

1. Anuidade anual do cartão (valor cheio)
2. O gasto de R$ {gasto}/mês isenta total ou parcialmente? (sim/não/parcial + condição exata)
3. Se não isenta: custo efetivo por mês (anuidade ÷ 12)
4. Retorno financeiro estimado com R$ {gasto}/mês:
   - Pontos/milhas acumulados por ano (use a taxa de pontos da base)
   - Valor estimado das milhas (use ~R$ 0,03/milha como referência)
   - Economia em lounge se aplicável (R$ 80-150 por acesso × acessos/ano)
   - Cashback direto se aplicável
5. Veredicto final: VALE ✅ ou NÃO VALE ❌
   - Se vale: com qual justificativa objetiva?
   - Se não vale: o que o usuário deveria fazer (trocar para qual alternativa?)"""

    claude(client, sistema, prompt)


def op_explicar(client: anthropic.Anthropic, sistema: str) -> None:
    hr()
    print("  EXPLICAR CARTÃO")
    hr()

    cartao = perguntar("\n  Nome do cartão: ")

    prompt = f"""Explique o cartão "{cartao}" de forma completa e direta:

1. Para quem é indicado (perfil ideal)
2. Anuidade — valor e TODAS as formas de isentar
3. Programa de pontos/milhas/cashback — taxa e parceiros
4. Benefícios de lounge/viagem — detalhes (quantos acessos, qual rede)
5. Pré-requisito de renda ou investimento para aprovação
6. Principais vantagens e desvantagens (seja honesto)
7. Alguma mudança importante em 2026?

Se o cartão não estiver na base, informe e sugira o mais parecido disponível."""

    claude(client, sistema, prompt)


# ── Loop principal ──────────────────────────────────────────────────────────


def menu_principal() -> None:
    print()
    hr("═")
    print("   AGENTE DE CARTÕES DE CRÉDITO BRASIL 2026")
    hr("═")
    print("  [1]  Recomendar cartão para meu perfil")
    print("  [2]  Comparar dois cartões")
    print("  [3]  Calcular se vale pagar anuidade")
    print("  [4]  Explicar um cartão específico")
    print("  [0]  Sair")
    hr("═")


def main() -> None:
    base = os.path.dirname(os.path.abspath(__file__))
    docx_path = os.path.join(base, "referencia_cartoes_agente_2026.docx")
    xlsx_path = os.path.join(base, "cartoes_credito_brasil_2026_v2.xlsx")

    for path in (docx_path, xlsx_path):
        if not os.path.exists(path):
            print(f"Erro: arquivo não encontrado: {path}")
            sys.exit(1)

    print("Carregando base de conhecimento...", end="", flush=True)
    sistema = montar_sistema(ler_docx(docx_path), ler_xlsx(xlsx_path))
    print(" pronto.\n")

    client = anthropic.Anthropic()  # lê ANTHROPIC_API_KEY do ambiente

    acoes = {
        "1": op_recomendar,
        "2": op_comparar,
        "3": op_vale_anuidade,
        "4": op_explicar,
    }

    while True:
        menu_principal()
        escolha = perguntar("  Escolha uma opção: ")

        if escolha == "0":
            print("\n  Até logo!\n")
            break
        elif escolha in acoes:
            acoes[escolha](client, sistema)
            perguntar("  [Enter para voltar ao menu]")
        else:
            print("\n  Opção inválida.\n")


if __name__ == "__main__":
    main()
